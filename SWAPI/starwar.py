"""
Created on 2018.12.14
@sysuxwh

功能：爬取SWAPI所有关于Star Wars的数据
网站：https://swapi.co/ 
实现：使用selenium库，结合chrome浏览器，使用xpath获取节点信息，最后使用openpyxl库生成xlsx文件
"""

import time
import os
import sys

# import xlwt
from openpyxl.workbook import Workbook  
from openpyxl.utils import get_column_letter 

from selenium import webdriver
from selenium.webdriver.common.keys import Keys        
import selenium.webdriver.support.ui as ui        
from selenium.webdriver.common.action_chains import ActionChains


browser = webdriver.Chrome()
base_url = "https://swapi.co/"

def initOutfile():
    # global outfile
    # outfile = xlwt.Workbook(encoding = 'urf-8')
    global wbook
    wbook = Workbook()

def initXls(key):
    name = ['数据类型', '序号', '数据']
    # global outfile
    global sheet
    global row
    global wbook

    # sheet = outfile.add_sheet(key)
    sheet = wbook.active
    if sheet.title != 'Sheet':
        sheet = wbook.create_sheet()
    sheet.title = key
    # row = 0
    # for col in range(len(name)):
        # sheet.write(row, col, name[col])
    # row = row + 1

    # sheet['A1'] = name[0]
    # sheet['B1'] = name[1]
    # sheet['C1'] = name[2]
    # row = 2
    row = 1
    # outfile.save("./StarWars.xls")
    wbook.save("./StarWars.xlsx")

def writeXls(key, num, text):
    # global outfile
    global wbook
    global sheet
    global row
    # sheet.write(row, 0, key)
    # sheet.write(row, 1, num)
    # sheet.write(row, 2, text)
    sheet['A%s'%(row)] = num
    sheet['B%s'%(row)] = text
    row = row + 1
    # outfile.save("./StarWars.xls")
    wbook.save("./StarWars.xlsx")


def GetSearchContent(key, num):
    initXls(key)
    count = 0
    i = 0
    while i < num:
        count = count + 1
        browser.get(base_url + 'api/'+ key + '/' + str(count) + '/?format=json')
        text = browser.find_element_by_xpath("//pre").text
        if text != '{"detail":"Not found"}':
            i = i + 1
            print('Get data of ' + key + ' ' + str(count) + ' (' + str(i) + '/' + str(num) + ')')
            writeXls(key, count, text)     

if __name__ == '__main__':
    browser.get(base_url)

    count = {'planets':61, 'starships':37, 'vehicles':39, 'people':87, 'films':7, 'species':37}

    initOutfile()

    # while True:
    #     key = input('请输入爬取的数据名字（planets/starships/vehicles/people/films/species）：')
    #     if key in count:
    #         GetSearchContent(key, count[key])
    #     elif key == 'exit':
    #         os._exit(0)
    #     else:
    #         print('该接口不存在')

    for key in count:
        GetSearchContent(key, count[key])
    print('爬取结束')
    time.sleep(10)